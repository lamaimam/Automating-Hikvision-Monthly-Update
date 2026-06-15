"""
verify_engine.py - Read-only diagnostic for the attendance engine.

For a given date range, prints exactly what the engine sees and decides
for each (employee, day) pair. No writes to Zoho. No emails. No state changes.

Use:
    python verify_engine.py --start 2026-05-01 --end 2026-05-21
    python verify_engine.py --start 2026-05-01 --end 2026-05-21 --employee 231
    python verify_engine.py --start 2026-05-15 --end 2026-05-15  # one day only
    python verify_engine.py --start 2026-05-01 --end 2026-05-21 --csv out.csv

Output columns:
    date | emp_id | name | hik_in | hik_out | duration | leave_type | rule | status | OT_hours

The "rule" column shows which constraint branch was applied:
    R1_no_event_no_leave_ASSUMED_PRESENT
    R1_no_event_annual_leave
    R1_no_event_sick_leave
    R1_no_event_other_leave
    R2_checkin_no_checkout_STANDARD_9H_NO_OT
    R3_under_9h_NO_OT
    R3_over_9h_WITH_OT
    R6_leave_and_checkin_NEEDS_REVIEW

This lets you eyeball any row: do the inputs (hik_in/hik_out/leave_type)
match what you see in iVMS-4200 and Zoho People for that day? Does the
applied rule match what you'd expect? Does OT_hours match the rule output?

Each row is one (employee, day) decision. Scan them for anomalies.
"""
import argparse
import csv
import logging
import sys
from datetime import date

import config
from attendance_engine import reconcile_day
from hikvision_client import HikvisionClient
from zoho_people_client import ZohoPeopleClient
from zoho_sheet_client import ZohoSheetClient


def setup_logging(verbose: bool) -> None:
    logging.basicConfig(
        level=logging.DEBUG if verbose else logging.WARNING,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        stream=sys.stderr,
    )


def classify_rule(shift, leave_cat, day_result) -> str:
    """Return a string label identifying which constraint branch fired."""
    has_in = bool(shift and shift.check_in)
    has_out = bool(shift and shift.check_out)
    on_leave = leave_cat is not None

    if on_leave and has_in:
        return "R6_leave_and_checkin_NEEDS_REVIEW"
    if not has_in and not has_out:
        if leave_cat == "annual":
            return "R1_no_event_annual_leave"
        if leave_cat == "sick":
            return "R1_no_event_sick_leave"
        if leave_cat == "other":
            return "R1_no_event_other_leave"
        return "R1_no_event_no_leave_ASSUMED_PRESENT"
    if has_in and not has_out:
        return "R2_checkin_no_checkout_STANDARD_9H_NO_OT"
    # Both in and out
    duration = shift.duration_hours if shift else None
    if duration is None:
        return "R_unknown_both_in_out_no_duration"
    if duration < config.STANDARD_WORKDAY_HOURS:
        return f"R3_under_9h_NO_OT (worked {duration:.2f}h)"
    return f"R3_over_9h_WITH_OT (worked {duration:.2f}h)"


def fmt_time(dt) -> str:
    return dt.strftime("%H:%M") if dt else "-"


def fmt_duration(shift) -> str:
    if not shift:
        return "-"
    d = shift.duration_hours
    return f"{d:.2f}h" if d is not None else "-"


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Verify attendance engine decisions for a date range."
    )
    parser.add_argument("--start", required=True, help="ISO date, e.g. 2026-05-01")
    parser.add_argument("--end", required=True, help="ISO date, e.g. 2026-05-21")
    parser.add_argument(
        "--employee",
        help="Filter to one employee ID (e.g. 231). Default: all employees "
             "in the contracts tab matching --end's month.",
    )
    parser.add_argument(
        "--csv",
        help="Write output to this CSV file instead of printing to stdout.",
    )
    parser.add_argument(
        "--xlsx-stacked",
        help="Write to an .xlsx with all employees on one sheet, "
             "each as a separate table block stacked vertically.",
    )
    parser.add_argument(
        "-v", "--verbose", action="store_true",
        help="Show DEBUG logs from clients (Hikvision pagination, etc.)",
    )
    args = parser.parse_args()

    setup_logging(args.verbose)

    start = date.fromisoformat(args.start)
    end = date.fromisoformat(args.end)
    if end < start:
        print("ERROR: --end is before --start", file=sys.stderr)
        return 2

    print(f"Fetching data for {start} -> {end}...", file=sys.stderr)

    sheet_client = ZohoSheetClient()
    contracts = sheet_client.read_contracts(end.year, end.month)
    if args.employee:
        contracts = [c for c in contracts if c.employee_id == args.employee]
        if not contracts:
            print(f"ERROR: employee {args.employee} not found in contracts "
                  f"for {end.year}-{end.month:02d}", file=sys.stderr)
            return 1

    print(f"Loaded {len(contracts)} contract(s)", file=sys.stderr)

    hik = HikvisionClient()
    shifts = hik.get_shifts(start, end)
    print(f"Hikvision: {sum(len(d) for d in shifts.values())} employee-days "
          f"across {len(shifts)} employee(s)", file=sys.stderr)

    people = ZohoPeopleClient()
    leaves = people.get_approved_leaves(start, end)
    print(f"Zoho People: {sum(len(d) for d in leaves.values())} approved "
          f"leave-days across {len(leaves)} employee(s)", file=sys.stderr)

    # Build the iteration: each (employee, day) in the window
    rows = []
    cur = start
    while cur <= end:
        for c in contracts:
            shift = shifts.get(c.employee_id, {}).get(cur)
            leave_cat = leaves.get(c.employee_id, {}).get(cur)
            day_result = reconcile_day(cur, shift, leave_cat)
            rule = classify_rule(shift, leave_cat, day_result)
            rows.append({
                "date": cur.isoformat(),
                "emp_id": c.employee_id,
                "name": c.name,
                "hik_in": fmt_time(shift.check_in if shift else None),
                "hik_out": fmt_time(shift.check_out if shift else None),
                "duration": fmt_duration(shift),
                "leave_type": leave_cat or "-",
                "rule": rule,
                "status": day_result.status,
                "OT_hours": f"{day_result.overtime_hours:.2f}",
                "notes": day_result.notes,
            })
        cur = date.fromordinal(cur.toordinal() + 1)

    fields = ["date", "emp_id", "name", "hik_in", "hik_out", "duration",
              "leave_type", "rule", "status", "OT_hours", "notes"]

    if args.csv:
        with open(args.csv, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fields)
            w.writeheader()
            w.writerows(rows)
        print(f"\nWrote {len(rows)} rows to {args.csv}", file=sys.stderr)
    elif args.xlsx_stacked:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print(
                "ERROR: openpyxl is not installed. Run:\n"
                "  pip install openpyxl",
                file=sys.stderr,
            )
            return 1

        wb = Workbook()
        ws = wb.active
        ws.title = "Verification"

        # Group rows by employee (preserving contract order = the order they
        # appeared in the contracts sheet).
        rows_by_emp: Dict[str, List[dict]] = {}
        emp_name: Dict[str, str] = {}
        for r in rows:
            rows_by_emp.setdefault(r["emp_id"], []).append(r)
            emp_name[r["emp_id"]] = r["name"]

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2C3E50")
        emp_title_font = Font(bold=True, size=14, color="2C3E50")
        center = Alignment(horizontal="center")

        # Highlight fills for status visibility
        leave_fill = PatternFill("solid", fgColor="FFF3CD")    # pale yellow
        review_fill = PatternFill("solid", fgColor="F8D7DA")   # pale red
        ot_fill = PatternFill("solid", fgColor="D4EDDA")       # pale green

        cur_row = 1
        for emp_id, emp_rows in rows_by_emp.items():
            # Employee title row
            ws.cell(row=cur_row, column=1,
                    value=f"{emp_name[emp_id]} ({emp_id})").font = emp_title_font
            ws.merge_cells(
                start_row=cur_row, start_column=1,
                end_row=cur_row, end_column=len(fields),
            )
            cur_row += 1

            # Header row
            for col_idx, fld in enumerate(fields, start=1):
                c = ws.cell(row=cur_row, column=col_idx, value=fld)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center
            cur_row += 1

            # Data rows
            for r in emp_rows:
                for col_idx, fld in enumerate(fields, start=1):
                    c = ws.cell(row=cur_row, column=col_idx, value=r[fld])
                # Highlight based on status
                status = r["status"]
                if status in ("annual_leave", "sick_leave", "absent_excused"):
                    fill = leave_fill
                elif status == "needs_review":
                    fill = review_fill
                elif status == "present_with_ot":
                    fill = ot_fill
                else:
                    fill = None
                if fill:
                    for col_idx in range(1, len(fields) + 1):
                        ws.cell(row=cur_row, column=col_idx).fill = fill
                cur_row += 1

            # Blank separator row between employees
            cur_row += 1

        # Auto-size columns based on content
        from openpyxl.utils import get_column_letter
        for col_idx, fld in enumerate(fields, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(fld)
            for r in rows:
                v = str(r[fld])
                if len(v) > max_len:
                    max_len = len(v)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        wb.save(args.xlsx_stacked)
        print(
            f"\nWrote {len(rows)} rows for {len(rows_by_emp)} employee(s) "
            f"to {args.xlsx_stacked}",
            file=sys.stderr,
        )
    else:
        # Pretty-print table to stdout
        widths = {fld: max(len(fld), max((len(str(r[fld])) for r in rows), default=0))
                  for fld in fields}
        sep = "  "
        # Header
        print(sep.join(f"{fld:<{widths[fld]}}" for fld in fields))
        print(sep.join("-" * widths[fld] for fld in fields))
        for r in rows:
            print(sep.join(f"{str(r[fld]):<{widths[fld]}}" for fld in fields))

    # Summary at end
    summary = {}
    for r in rows:
        summary[r["rule"]] = summary.get(r["rule"], 0) + 1
    print("\n=== Rule application summary ===", file=sys.stderr)
    for rule, count in sorted(summary.items(), key=lambda x: -x[1]):
        print(f"  {count:4d}  {rule}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    sys.exit(main())